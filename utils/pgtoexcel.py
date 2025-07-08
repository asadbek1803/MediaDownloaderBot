import openpyxl
from openpyxl.styles import Font
from datetime import datetime

async def export_to_excel(data, headings, filepath):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.row_dimensions[1].font = Font(bold=True)

    # Yozuvlar sarlavhasi
    for colno, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=colno).value = heading

    # Asosiy ma'lumotlar
    for rowno, row in enumerate(data, start=2):
        for colno, cell_value in enumerate(row, start=1):
            # Agar datetime bo'lsa va tzinfo bor bo'lsa — uni tozalaymiz
            if isinstance(cell_value, datetime) and cell_value.tzinfo is not None:
                cell_value = cell_value.replace(tzinfo=None)

            sheet.cell(row=rowno, column=colno).value = cell_value

    wb.save(filepath)
